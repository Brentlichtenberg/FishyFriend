#!/usr/bin/env python3
"""
Extract with bold/regular text formatting preserved using openpyxl for Excel output.
"""

import pdfplumber
import re
from openpyxl import Workbook
from openpyxl.styles import Font

def extract_text_with_formatting(chars):
    """
    Extract text and determine which words are bold.
    Returns: (text_string, bold_positions) where bold_positions is list of (start, end) tuples
    """
    text = ""
    bold_ranges = []
    current_pos = 0
    
    for char in chars:
        char_text = char.get('text', '')
        font = char.get('fontname', '')
        is_bold = 'Bd' in font
        
        if char_text:
            if is_bold:
                # Track bold character positions
                start = current_pos
                end = current_pos + len(char_text)
                # Merge with previous range if adjacent
                if bold_ranges and bold_ranges[-1][1] == start:
                    bold_ranges[-1] = (bold_ranges[-1][0], end)
                else:
                    bold_ranges.append((start, end))
            
            text += char_text
            current_pos += len(char_text)
    
    return text, bold_ranges

def is_text_bold(text, page_chars, line_start_pos):
    """Check if a specific text segment is bold in the PDF."""
    # Find the characters for this text
    chars_in_range = []
    char_pos = 0
    
    for char in page_chars:
        if char_pos >= line_start_pos and char_pos < line_start_pos + len(text):
            chars_in_range.append(char)
        char_pos += len(char.get('text', ''))
    
    # Count bold vs regular characters
    bold_count = sum(1 for c in chars_in_range if 'Bd' in c.get('fontname', ''))
    
    # If more than 50% of characters are bold, consider the text bold
    return bold_count > len(chars_in_range) / 2 if chars_in_range else False

def extract_bathroom_products_with_formatting(pdf_path, start_page, end_page, output_path):
    """Extract with formatting preserved in Excel."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Set column widths
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    
    # Add headers
    headers = ['Section', 'Net Wt', 'SKU', 'Color', 'Price', 'QR Code URL']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    row_num = 2
    
    with pdfplumber.open(pdf_path) as pdf:
        for page_num in range(start_page - 1, end_page):
            page = pdf.pages[page_num]
            actual_page_num = page_num + 1
            print(f"\nProcessing page {actual_page_num}...")
            
            text = page.extract_text()
            if not text:
                continue
            
            # Get character-level data for formatting
            page_chars = page.chars
            
            lines = text.split('\n')
            
            current_product_title = ""
            current_nominal_dimensions = ""
            product_title_added = False
            in_description_section = False
            in_separate_components = False
            in_required_components = False
            
            for i, line in enumerate(lines):
                line = line.strip()
                
                if not line or 'americanstandard-us.com' in line.lower() or 'list price guide' in line.lower():
                    continue
                
                skip_lines = ['Bathroom Furniture', 'Furniture', 'Vanities', 'Bathroom Sinks', 
                             'Bathroom', 'Sinks', 'Pedestal Sinks', 'Pedestal', 'Above Counter',
                             'Drop-In', 'Under Counter', 'Wall-Hung', 'seitinaV', 'erutinruF',
                             'skniS', 'moorhtaB', 'latsedeP', 'Bowl Size:', 'Parts Program']
                if line in skip_lines or line.startswith('Bowl Size:'):
                    continue
                
                # Detect product title
                is_title = (('®' in line or '™' in line) and 
                           not line.startswith('•') and 
                           not line.startswith('Nominal Dimensions') and
                           not line.startswith('Description') and
                           not line.startswith('Separate Components') and
                           not line.startswith('Required Components') and
                           not re.search(r'\b[0-9]{4,10}[.-][0-9A-Z.-]+\b', line) and
                           'Net Wt' not in line)
                
                if is_title:
                    current_product_title = line
                    product_title_added = False
                    in_description_section = False
                    in_separate_components = False
                    in_required_components = False
                    print(f"  Title: {current_product_title[:60]}...")
                    continue
                
                if line.startswith('Nominal Dimensions'):
                    if i + 1 < len(lines):
                        current_nominal_dimensions = lines[i + 1].strip()
                    continue
                
                if 'Description' in line and 'Net Wt' in line and 'SKU' in line:
                    in_description_section = True
                    in_separate_components = False
                    in_required_components = False
                    continue
                
                if line == 'Separate Components:':
                    in_description_section = False
                    in_separate_components = True
                    in_required_components = False
                    continue
                
                if line == 'Required Components:':
                    in_description_section = False
                    in_separate_components = False
                    in_required_components = True
                    continue
                
                # Extract product data rows
                if in_description_section or in_separate_components or in_required_components:
                    sku_match = re.search(r'\b([0-9]{4,10}[.-][0-9A-Z.-]+)\b', line)
                    
                    if sku_match:
                        sku = sku_match.group(1)
                        
                        price_match = re.search(r'\b(\d{1,2},?\d{3})\s*$', line)
                        if not price_match:
                            price_match = re.search(r'\b(\d{2,4})\s*$', line)
                        price = price_match.group(1) if price_match else ""
                        
                        color = ""
                        if price_match:
                            text_between = line[sku_match.end():price_match.start()].strip()
                            color = text_between
                        
                        weight = ""
                        weight_match = re.search(r'(\d+(?:\.\d+)?\s*lb)', line[:sku_match.start()])
                        if weight_match:
                            weight = weight_match.group(1)
                        elif i > 0:
                            prev_line = lines[i-1].strip()
                            weight_match = re.search(r'(\d+(?:\.\d+)?\s*kg)', prev_line)
                            if weight_match:
                                weight = weight_match.group(1)
                        
                        description = ""
                        if weight_match:
                            description = line[:weight_match.start()].strip()
                        else:
                            description = line[:sku_match.start()].strip()
                        
                        if len(description) < 5 and i > 0:
                            prev_line = lines[i-1].strip()
                            if (prev_line and 
                                not prev_line.startswith('•') and
                                not prev_line.startswith('Nominal') and
                                not prev_line.endswith(':') and
                                not re.search(r'^\d+\.?\d*\s*kg\s*$', prev_line)):
                                description = prev_line
                        
                        # Add product title section if first Description entry
                        if in_description_section and not product_title_added and current_product_title:
                            # Product title row (bold)
                            cell = ws.cell(row=row_num, column=1, value=current_product_title)
                            cell.font = Font(bold=True)
                            row_num += 1
                            
                            # Empty row
                            row_num += 1
                            
                            # Nominal Dimensions
                            if current_nominal_dimensions:
                                cell = ws.cell(row=row_num, column=1, value='Nominal Dimensions')
                                cell.font = Font(bold=True)
                                row_num += 1
                                
                                ws.cell(row=row_num, column=1, value=current_nominal_dimensions)
                                row_num += 1
                            
                            # Description header
                            cell = ws.cell(row=row_num, column=1, value='Description')
                            cell.font = Font(bold=True)
                            row_num += 1
                            
                            product_title_added = True
                        
                        # Add section header for Separate/Required Components
                        if in_separate_components:
                            if row_num == 2 or ws.cell(row=row_num-1, column=1).value != 'Separate Components':
                                row_num += 1  # Empty row
                                cell = ws.cell(row=row_num, column=1, value='Separate Components')
                                cell.font = Font(bold=True)
                                row_num += 1
                        
                        if in_required_components:
                            if row_num == 2 or ws.cell(row=row_num-1, column=1).value != 'Required Components':
                                row_num += 1  # Empty row
                                cell = ws.cell(row=row_num, column=1, value='Required Components')
                                cell.font = Font(bold=True)
                                row_num += 1
                        
                        # Add product row with regular font
                        ws.cell(row=row_num, column=1, value=description)
                        ws.cell(row=row_num, column=2, value=weight)
                        ws.cell(row=row_num, column=3, value=sku)
                        ws.cell(row=row_num, column=4, value=color)
                        ws.cell(row=row_num, column=5, value=price)
                        ws.cell(row=row_num, column=6, value='')
                        
                        row_num += 1
                        print(f"    {sku} - {description[:40]}... | ${price}")
    
    wb.save(output_path)
    
    print(f"\n{'='*70}")
    print(f"✓ Extracted {row_num - 2} entries from pages {start_page}-{end_page}")
    print(f"✓ Saved to: {output_path}")
    print(f"{'='*70}")

if __name__ == "__main__":
    pdf_path = "/Users/brentlichtenberg/Desktop/PSKU Project/American Standard Residential Product Catalog.pdf"
    output_path = "/Users/brentlichtenberg/Desktop/PSKU Project/BRL Options-2026-Populated.xlsx"
    
    print("Extracting Bathroom Furniture and Bathroom Sinks with formatting...")
    print("Output will be in Excel format to preserve bold text formatting.")
    
    extract_bathroom_products_with_formatting(pdf_path, 100, 130, output_path)
